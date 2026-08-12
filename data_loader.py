"""
Parses Digital_readiness_tool_Framework.xlsx into plain Python structures.

The Excel file is the single source of truth for framework content, MVS
benchmarks/savings, the solutions bank, and customer stories — editing the
workbook (not this code) is how content owners update the tool.

Sheets:
    framework              -> FRAMEWORK[dimension] (levels, solutions, next step, MVS)
    MVS potential savings  -> MVS_SAVINGS[industry_sub_sector]
    solutions bank         -> SOLUTIONS_PORTFOLIO[product_name] -> portfolio label
    customer stories bank  -> CUSTOMER_STORIES (list of story dicts)
    profiles               -> PROFILES (list of (bucket_label, description))
"""

from __future__ import annotations

import functools
import os

import openpyxl
import streamlit as st

XLSX_PATH = os.path.join(os.path.dirname(__file__), "Digital_readiness_tool_Framework.xlsx")

DIMENSIONS: list[str] = ["strategy", "people", "operations", "connectivity", "intelligence"]

# The "framework" sheet spells out each dimension name slightly differently
# (extra whitespace, parenthetical subtitle) — map to our canonical ids.
_DIMENSION_NAME_MAP = {
    "strategy": "strategy",
    "people": "people",
    "operations (operational excellence)": "operations",
    "connectivity (connected factory)": "connectivity",
    "intelligence (intelligent factory)": "intelligence",
}

# Food categories offered on the participant form, mapped to the "Industry
# Sub-Sector" label used in the "MVS potential savings" sheet. The sheet has
# no entry for "Plant-based" or "Other": Plant-based beverages are commercially
# closest to non-dairy beverages, so they fall back to that row; "Other" has
# no sensible single match, so it falls back to the average across all five
# rows (computed in _average_savings_row below).
FOOD_CATEGORY_TO_SUBSECTOR = {
    "Dairy": "Dairy Beverage",
    "Beverage": "Beverages (Non-Dairy)",
    "Cheese": "Cheese & Whey",
    "Powder": "Powder / Reconstitution",
    "Ice Cream": "Ice Cream",
    "Plant-based": "Beverages (Non-Dairy)",
    "Other": "__AVERAGE__",
}


def _clean(v):
    return v.strip() if isinstance(v, str) else v


@st.cache_resource(show_spinner=False)
def load_workbook_data() -> dict:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    return {
        "framework": _parse_framework(wb["framework"]),
        "mvs_savings": _parse_mvs_savings(wb["MVS potential savings"]),
        "solutions_bank": _parse_solutions_bank(wb["solutions bank"]),
        "customer_stories": _parse_customer_stories(wb["customer stories bank"]),
        "profiles": _parse_profiles(wb["profiles"]),
    }


def _parse_framework(ws) -> dict:
    framework: dict = {}
    current_dim = None
    current_level = None

    for row in ws.iter_rows(min_row=7, max_row=ws.max_row):
        name_cell = _clean(row[7].value)   # H
        level_cell = row[8].value          # I
        level_name = _clean(row[9].value)  # J
        description = _clean(row[10].value)  # K
        solution = _clean(row[11].value)   # L
        value_prop = _clean(row[12].value)  # M
        mvs_cell = row[6].value            # G

        # A new dimension block starts whenever column H holds a recognized
        # dimension name — that happens on the "0/Not Assessed" row for
        # Strategy, but on the level-1 row for every other dimension (the
        # sheet isn't laid out consistently), so we key off the name itself
        # rather than a fixed level number.
        dim_key = _DIMENSION_NAME_MAP.get((name_cell or "").lower().strip())
        if dim_key is not None:
            current_dim = dim_key
            current_level = level_cell if isinstance(level_cell, int) else 0
            framework[current_dim] = {
                "mvs": mvs_cell,
                "question": "",
                "levels": {},
                "solutions": {lvl: [] for lvl in range(1, 6)},
                "next_step": "",
            }
            if level_cell == 0:
                continue

        if current_dim is None:
            continue

        if isinstance(level_cell, int) and 1 <= level_cell <= 5:
            current_level = level_cell
            framework[current_dim]["levels"][level_cell] = {
                "name": level_name,
                "description": description,
            }
            if name_cell:
                framework[current_dim]["question"] = name_cell
            if solution:
                framework[current_dim]["solutions"][level_cell].append(
                    {"name": solution, "vp": value_prop}
                )
            continue

        if isinstance(level_cell, str) and level_cell.strip().lower().startswith("next step"):
            framework[current_dim]["next_step"] = solution or ""
            continue

        # Continuation row: extra solution(s) for the current level.
        if solution and current_level:
            framework[current_dim]["solutions"][current_level].append(
                {"name": solution, "vp": value_prop}
            )

    return framework


def _parse_pct_range(text: str) -> tuple[float, float, str] | None:
    """'+8% to +12%' -> (8.0, 12.0, '+')."""
    if not isinstance(text, str) or "to" not in text:
        return None
    sign = "-" if text.strip().startswith("-") else "+"
    parts = text.replace("%", "").replace("+", "").replace("-", "").split("to")
    try:
        lo, hi = float(parts[0].strip()), float(parts[1].strip())
    except (ValueError, IndexError):
        return None
    return lo, hi, sign


def _format_pct_range(lo: float, hi: float, sign: str) -> str:
    return f"{sign}{lo:.0f}% to {sign}{hi:.0f}%"


def _average_savings_row(rows: dict) -> dict:
    """Average each KPI's range across every parsed sub-sector row."""
    kpis = ["oee", "quality", "energy", "stock"]
    avg = {}
    for kpi in kpis:
        parsed = [_parse_pct_range(r[kpi]) for r in rows.values() if _parse_pct_range(r[kpi])]
        if not parsed:
            avg[kpi] = ""
            continue
        lo = sum(p[0] for p in parsed) / len(parsed)
        hi = sum(p[1] for p in parsed) / len(parsed)
        sign = parsed[0][2]
        avg[kpi] = _format_pct_range(lo, hi, sign)
    return avg


def _parse_mvs_savings(ws) -> dict:
    savings: dict = {}
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        subsector = _clean(row[1].value)  # B
        if not subsector:
            continue
        savings[subsector] = {
            "oee": _clean(row[2].value),
            "quality": _clean(row[3].value),
            "energy": _clean(row[4].value),
            "stock": _clean(row[5].value),
        }
    savings["__AVERAGE__"] = _average_savings_row(savings)
    return savings


def _parse_solutions_bank(ws) -> dict:
    bank: dict = {}
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        product = _clean(row[2].value)  # C
        vp = _clean(row[3].value)       # D
        portfolio = _clean(row[4].value)  # E
        if product:
            bank[product] = {"vp": vp, "portfolio": portfolio}
    return bank


_LEVEL_COLS = {"strategy": 4, "people": 5, "operations": 6, "connectivity": 7, "intelligence": 8}


def _bucket_for_cell(value) -> str | None:
    """Map a raw customer-stories-bank dimension cell to a maturity bucket:
    '1-2', '3', '4-5', or None if the cell doesn't imply a level (blank, or
    free-text unrelated to level like 'Low in connectivity' handled below)."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        lvl = int(value)
        if lvl <= 2:
            return "1-2"
        if lvl == 3:
            return "3"
        return "4-5"
    text = str(value).strip().lower()
    if "low" in text:
        return "1-2"
    if "4 and 5" in text or "high" in text:
        return "4-5"
    return None


def _parse_customer_stories(ws) -> list[dict]:
    stories = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        name = _clean(row[1].value)  # B
        if not name or name.strip().lower() == "logic":
            continue
        category = _clean(row[2].value)  # C
        url = _clean(row[9].value)  # J
        for dim, col in _LEVEL_COLS.items():
            raw = row[col].value
            bucket = _bucket_for_cell(raw)
            if bucket is None:
                continue
            stories.append(
                {
                    "customer": name,
                    "food_category": category or "",
                    "dimension": dim,
                    "maturity_bucket": bucket,
                    "url": url or "",
                }
            )
    return stories


def _parse_profiles(ws) -> list[tuple[str, str]]:
    profiles = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        label = _clean(row[0].value)
        desc = _clean(row[1].value)
        if label:
            profiles.append((label, desc or ""))
    return profiles


def savings_row_for_category(mvs_savings: dict, food_category: str) -> dict:
    subsector = FOOD_CATEGORY_TO_SUBSECTOR.get(food_category, "__AVERAGE__")
    return mvs_savings.get(subsector, mvs_savings["__AVERAGE__"])
